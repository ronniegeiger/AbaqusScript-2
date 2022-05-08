import datetime
import string
file_name = 'Data.txt'
with open(file_name,'r',encoding='utf-8') as DA:
    # Data=DA.read().strip()
    Data=DA.readlines()
    Pkeys=str.split(Data[0],' ')
pt1=[]
pt2=[]
for line in Data[2:]:
    ele=str.split(line,' ')
    pt1.append(ele[0])#Strain
    pt2.append(ele[1])#Stress
""" 
This module is used to write data from txt file to Excel
"""
# import the library at here to prevent error which "can't open'Data.txt' in xlsx format"
from openpyxl import *
# Create New Workbook
wb = Workbook()
# ws=wb.active
sheetname=wb.sheetnames[0]
ws=wb[sheetname]
# Write Titles
ws['A1']=Pkeys[0]
ws['B1']=Pkeys[1]
# Write data
# ws.cell(row=2,column=1).value=1
# print(ws.cell(row=2,column=1).value)
for i in range(2,len(Data)):
    # print(i)
    ws.cell(row=i,column=1).value=pt1[i-2]
    ws.cell(row=i,column=2).value=pt2[i-2]
wb.save("Data.xlsx")
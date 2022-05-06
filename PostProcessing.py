from abaqus import *
from abaqusConstants import *
from textRepr import *
import mesh
import time
import odbAccess
from odbAccess import *
import visualization
import regionToolset
import numpy as np
from openpyxl import *
import datetime

# Initialize Variables
length = 38.0	
width = 12.5
height_plate = 3.0
spacing_y=0.5
spacing_x = 0.5
num_DatumPlanes_y = int((width /spacing_y) -1.0)
num_DatumPlanes_x = int((length/spacing_x) -1.0)
num_plies = 11
ply=0
sectionpoint=3
# Architecture
omega=1.0
Firstphase=1.57

#postprocessing pre
odbfilename = mdb.jobs.keys()[0]
CrossSectionalArea= width * height_plate
# PostProcessing
odb = session.openOdb(name='myodb',path='{}.odb'.format(odbfilename))
Frames = odb.steps['Step-1'].frames
#
FOPSet=odb.rootAssembly.nodeSets['SET-TENSILE']
xydata=[]
for Frame in range(len(Frames)):
	RFarray=[]#create array and clear the array
	Uarray =[]#create array and clear the array
	RF=Frames[Frame].fieldOutputs['RF']
	U =Frames[Frame].fieldOutputs['U']
	RFsubset=RF.getSubset(region=FOPSet).values
	Usubset = U.getSubset(region=FOPSet).values
	for RFvalue in range(len(RFsubset)): 
		RFarray.append(RFsubset[RFvalue].data[0]) #data[0] means RF1
	for Uvalue in range(len(Usubset)): 
		Uarray.append(Usubset[Uvalue].data[0])    #data[0] means U1
	stress=np.mean(RFarray)/CrossSectionalArea
	strain=np.mean(Uarray)/(length*2)
	xydata.append((strain,stress))
""" 
# Export fundamental information
file_fundamentalinfo='FundamentalInfo'
# Write the fundamental information to txt file
with open('{}.txt'.format(file_fundamentalinfo),'w') as FI:
    FI.write("Initialize Variables:\n")
    FI.write('Length:{}\n'.format(length))
    FI.write('Width:{}\n'.format(width))
    FI.write('Height:{}\n'.format(height_plate))
    FI.write('Architecture Function:Z={A}sin({omega}x+{FirstPhase})+{z0}'.format(A=(height_plate/2),omega=omega,FirstPhase=Firstphase,z0=(height_plate/2)))
    # FI.write() 
"""
# """ 
# Export Data
file_data='Data'
with open('{}.txt'.format(file_data),'w') as FD:
    FD.write("Strain Stress\n")
	for i in range(len(xydata)):
		FD.write(xydata[i][0]) #strain
		FD.write(" ")
		FD.write(xydata[i][1]) #stress
	
    
# """
""" 
# Export Data
wb = Workbook()
sheetname=wb.sheetnames[0]
ws=wb[sheetname]
ws['A1']='Strain'
ws['B1']='Stress'
ws['E1']=datetime.datetime.now()
for i in range(len(xydata)):
    num=i+2
    sheetname['{}{}'.format('A',num)]=xydata[i][0]
    sheetname['{}{}'.format('B',num)]=xydata[i][1]
wb.save("data.xlsx")
"""